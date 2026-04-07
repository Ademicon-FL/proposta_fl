"""
leitor.py — Funções de leitura e extração dos dados da planilha Ademicon
"""
import io
import datetime
import openpyxl


def _norm(val):
    """Normalize cell value to lowercase stripped string for comparison."""
    if val is None:
        return ""
    return str(val).strip().lower()


def _fmt_mes(dt):
    """Format datetime to 'MM/YYYY' string."""
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%m/%Y")
    return str(dt)


# ──────────────────────────────────────────────────────────────────────────────
# RESUMO
# ──────────────────────────────────────────────────────────────────────────────
def ler_resumo(ws):
    """
    Extract from RESUMO sheet:
    - credito_total, qtd_parcelas, valor_parcela, taxa_estatica, tir
    - fluxo: list of {mes, parcela, credito, credito_acumulado}
    """
    dados = {}
    fluxo = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        # col B (index 1) is the label, col C (index 2) is the value
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        e = row[4] if len(row) > 4 else None
        bn = _norm(b)

        if "total crédito levantado" in bn or "total credito levantado" in bn:
            dados["credito_total"] = c
        elif "quantidade de parcelas" in bn:
            dados["qtd_parcelas"] = c
        elif "valor da parcela" in bn:
            dados["valor_parcela"] = c
        elif "taxa estática" in bn or "taxa estatica" in bn:
            dados["taxa_estatica"] = c
        elif bn == "tir":
            dados["tir_mensal"] = c
            if c is not None:
                dados["tir_anual"] = (1 + c) ** 12 - 1
        # Header row for monthly flow
        elif "mês" in bn and _norm(c) == "parcela" and not header_found:
            header_found = True
        # Data rows below the header
        elif header_found and isinstance(b, datetime.datetime):
            fluxo.append({
                "mes": b,
                "parcela": c,
                "credito": d,
                "credito_acumulado": e,
            })

    dados["fluxo"] = fluxo
    return dados


# ──────────────────────────────────────────────────────────────────────────────
# CARTEIRA
# ──────────────────────────────────────────────────────────────────────────────
def ler_carteira(ws):
    """
    Extract from CARTEIRA sheet:
    - grupos: list of group rows
    - totais: dict with totals per section + carteira total
    - data_proposta, credito_total_carteira, prazo_medio, pct_fixo
    """
    grupos = []
    totais = {}
    header_found = False
    secao_atual = "LANCE LIVRE"
    data_proposta = None
    credito_total_carteira = None
    prazo_medio = None
    pct_fixo = None

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        # Date in col D (index 3) — very first data rows
        if data_proposta is None and isinstance(row[3] if len(row) > 3 else None, datetime.datetime):
            data_proposta = row[3]

        row_text = " ".join(_norm(v) for v in row if v is not None)

        # Detect the header row by looking for "crédito contratado"
        if not header_found and ("crédito contratado" in row_text or "credito contratado" in row_text):
            header_found = True
            continue

        if not header_found:
            continue

        b = row[1] if len(row) > 1 else None  # section label
        c = row[2] if len(row) > 2 else None  # group number or total label
        bn = _norm(b)
        cn = _norm(c)

        # Section label
        if bn in ("lance livre",):
            secao_atual = "LANCE LIVRE"
        elif "fixo" in bn or "limitado" in bn:
            secao_atual = "FIXO/LIMITADO"

        # Total rows
        if c is not None and ("total lance livre" in cn or "total lance fixo" in cn or "total carteira" in cn):
            totais[str(c).strip()] = {
                "credito_contratado": row[3] if len(row) > 3 else None,
                "lance_embutido": row[6] if len(row) > 6 else None,
                "lance_livre": row[7] if len(row) > 7 else None,
                "qtde_cotas": row[8] if len(row) > 8 else None,
            }
            # TOTAL CARTEIRA also has aggregated info in right columns
            if "total carteira" in cn:
                credito_total_carteira = row[15] if len(row) > 15 else None
                prazo_medio = row[17] if len(row) > 17 else None
            continue

        # Percentage line
        if "% de fixo" in row_text or "% fixo" in row_text:
            pct_fixo = row[15] if len(row) > 15 else None
            continue

        # Group data rows: col C must be a number
        try:
            grupo_num = int(float(str(c).strip())) if c is not None else None
        except (ValueError, TypeError):
            grupo_num = None

        if grupo_num is not None:
            grupos.append({
                "secao": secao_atual,
                "grupo": grupo_num,
                "credito_contratado": row[3] if len(row) > 3 else None,
                "parcelas_pre": row[4] if len(row) > 4 else None,
                "prazo": row[5] if len(row) > 5 else None,
                "lance_embutido": row[6] if len(row) > 6 else None,
                "lance_livre": row[7] if len(row) > 7 else None,
                "qtde_cotas": row[8] if len(row) > 8 else None,
                "credito_novo": row[9] if len(row) > 9 else None,
            })

    return {
        "grupos": grupos,
        "totais": totais,
        "data_proposta": data_proposta,
        "credito_total": credito_total_carteira,
        "prazo_medio": prazo_medio,
        "pct_fixo": pct_fixo,
    }


# ──────────────────────────────────────────────────────────────────────────────
# FLUXO
# ──────────────────────────────────────────────────────────────────────────────
def ler_fluxo(ws):
    """
    Extract from FLUXO sheet:
    - credito_total, parcela, total_cotas, taxa_fidc, tir_mensal
    - fluxo: list of monthly rows, with 'contemplado' flag
    """
    dados = {}
    fluxo = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        b = row[1] if len(row) > 1 else None
        d = row[3] if len(row) > 3 else None
        bn = _norm(b)

        # Premissas
        if "crédito total contratado" in bn or "credito total contratado" in bn:
            dados["credito_total"] = d
        elif "parcela pré" in bn or "parcela pre" in bn:
            dados["parcela"] = d
        elif "total de cotas" in bn and "grupo" not in bn:
            dados["total_cotas"] = d
        elif "percentual lance embutido" in bn:
            dados["pct_lance_embutido"] = d

        # FIDC rate in col G (index 6)
        if len(row) > 6 and row[6] is not None and isinstance(row[6], (int, float)) and 0.05 < row[6] < 1.0:
            if "taxa" in _norm(row[5] if len(row) > 5 else None) or "tx" in _norm(row[5] if len(row) > 5 else None):
                dados["taxa_fidc"] = row[6]
            elif row[6] == 0.24:  # known default FIDC rate
                dados["taxa_fidc"] = row[6]

        # Header row: contains both 'cotas contempladas' and 'valor pago'
        row_text = " ".join(_norm(v) for v in row if v is not None)
        if not header_found and "cotas contempladas" in row_text and "valor pago" in row_text:
            header_found = True
            # TIR is embedded in the header row at col M (index 12)
            if len(row) > 12 and isinstance(row[12], float):
                dados["tir_mensal"] = row[12]
                dados["tir_anual"] = (1 + row[12]) ** 12 - 1
            continue

        if header_found and isinstance(row[0], datetime.datetime):
            cotas = row[2] if len(row) > 2 else None
            contemplado = cotas is not None and isinstance(cotas, (int, float)) and cotas > 0

            fluxo.append({
                "mes": row[0],
                "cotas_contempladas": cotas,
                "valor_pago": row[3] if len(row) > 3 else None,
                "valores_pagos_acumulados": row[4] if len(row) > 4 else None,
                "lance_pago": row[5] if len(row) > 5 else None,
                "lance_embutido": row[6] if len(row) > 6 else None,
                "credito_contemplado": row[7] if len(row) > 7 else None,
                "credito_liberado": row[8] if len(row) > 8 else None,
                "credito_liquido_pos_fidc": row[9] if len(row) > 9 else None,
                "credito_liquido_acumulado": row[10] if len(row) > 10 else None,
                "contemplado": contemplado,
            })

    dados["fluxo"] = fluxo
    return dados


# ──────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────
def ler_planilha(file_bytes_or_obj):
    """
    Read an Ademicon structured operation spreadsheet.

    Returns a dict with keys:
      sheets, resumo, carteira, fluxo, fluxo_sheet_name, erros
    """
    if isinstance(file_bytes_or_obj, bytes):
        obj = io.BytesIO(file_bytes_or_obj)
    else:
        obj = file_bytes_or_obj

    try:
        wb = openpyxl.load_workbook(obj, data_only=True)
    except Exception as e:
        return {"erros": [f"Não foi possível abrir o arquivo: {e}"], "sheets": []}

    sheets = wb.sheetnames
    resultado = {
        "sheets": sheets,
        "resumo": None,
        "carteira": None,
        "fluxo": None,
        "fluxo_sheet_name": None,
        "erros": [],
    }

    # RESUMO
    resumo_sheet = next((s for s in sheets if s.strip().upper() == "RESUMO"), None)
    if resumo_sheet:
        try:
            resultado["resumo"] = ler_resumo(wb[resumo_sheet])
        except Exception as e:
            resultado["erros"].append(f"Erro ao ler aba RESUMO: {e}")
    else:
        resultado["erros"].append("Aba 'RESUMO' não encontrada")

    # CARTEIRA
    carteira_sheet = next((s for s in sheets if s.strip().upper() == "CARTEIRA"), None)
    if carteira_sheet:
        try:
            resultado["carteira"] = ler_carteira(wb[carteira_sheet])
        except Exception as e:
            resultado["erros"].append(f"Erro ao ler aba CARTEIRA: {e}")
    else:
        resultado["erros"].append("Aba 'CARTEIRA' não encontrada")

    # FLUXO (any sheet with 'FLUXO' in the name)
    fluxo_sheets = [s for s in sheets if "FLUXO" in s.upper()]
    if fluxo_sheets:
        # Prefer COM FIDC if present
        com_fidc = [s for s in fluxo_sheets if "COM FIDC" in s.upper()]
        chosen = com_fidc[0] if com_fidc else fluxo_sheets[0]
        try:
            resultado["fluxo"] = ler_fluxo(wb[chosen])
            resultado["fluxo_sheet_name"] = chosen
        except Exception as e:
            resultado["erros"].append(f"Erro ao ler aba FLUXO: {e}")
    else:
        resultado["erros"].append("Aba 'FLUXO' não encontrada")

    return resultado
